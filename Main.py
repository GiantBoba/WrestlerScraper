import json
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from getData import *

print("Fire Pro Wrestling World - Wrestler Data Populator (0.1)")
print("--------------------------------------------------------")

#Import variables from Config.json file
def import_variables(config_file):
    with open(config_file, 'r') as file:
        variables = json.load(file)
    return variables
    
config_file = "Config.json"

# Import variables from the file
config = import_variables(config_file)
xlsx_file = config['xls file']
sheet_name = config['sheet name']
weight_limit = int(config['junior weight limit'])
super_m = config["superhuman m"]
legend_m = config["legend m"]
main_m = config["mainevent m"]
mid_m= config["midcard m"]
super_f = config["superhuman f"]
legend_f = config["legend f"]
main_f = config["mainevent f"]
mid_f = config["midcard f"]
s_size = int(config["small size"])
m_size = int(config["medium size"])
l_size = int(config["large size"])

wb = load_workbook(xlsx_file)
ws = wb[sheet_name]
all_rows = list(ws.rows)
sheet = wb.active
print("Sheet '" + sheet_name + "' successfully loaded from '" + xlsx_file + "' file.")
wb.save(filename = xlsx_file)
    
sheet["A1"] = "Full Name"
sheet["B1"] = "Nickname"
sheet["C1"] = "Parameter"
sheet["D1"] = "Gender"
sheet["E1"] = "Rank"
sheet["F1"] = "Active"
sheet["G1"] = "Promotion"
sheet["H1"] = "Stable"
sheet["I1"] = "Costumes"
sheet["J1"] = "Weight"
sheet["K1"] = "Weight Class"
sheet["L1"] = "Height"
sheet["M1"] = "Body Size"
sheet["N1"] = "Country"
sheet["O1"] = "Date of Birth"
sheet["P1"] = "Finisher"
sheet["Q1"] = "Finisher Description"
print("Headers populated")


for i in range(2, sheet.max_row+1):
#for i in range(2, 10):
    wrestler_name = sheet["A"+str(i)].value
    if(wrestler_name != None):
        print("Fetching data for: " + wrestler_name)
        response = getSource(wrestler_name)
        payload = BeautifulSoup(response.text, 'html.parser')
        sheet["B" + str(i)] = getNickname(payload)
        sheet["C" + str(i)] = 0
        gender = getGender(payload)
        sheet["D" + str(i)] = gender
        sheet["E" + str(i)] = '=IF(B' + str(i) + '="M", IF(C' + str(i) + '>' + super_m + ', "Super Human", IF(C' + str(i) + '>' + legend_m + ', "Legend", IF(C' + str(i) + '>' + main_m + ', "Main Event", IF(C' + str(i) + '>110, ' + mid_m + ', "Jobber")))), IF(C' + str(i) + '>' + super_f + ', "Super Human", IF(C' + str(i) + '>' + legend_f + ', "Legend", IF(C' + str(i) + '>' + main_f + ', "Main Event", IF(C' + str(i) + '>' + mid_f + ', "Mid-Carder", "Jobber")))))'
        isActive = getIsActive(payload)
        sheet["F" + str(i)] = isActive
        if(isActive == "Yes"):
            sheet["G" + str(i)] = getActivePromotion(payload)
            sheet["H" + str(i)] = getStable(payload)
        else:
            sheet["G" + str(i)] = getPromotion(payload)
            sheet["H" + str(i)] = "Legend"
        sheet["I" + str(i)] = 0
        sheet["J" + str(i)] = getWeight(payload)
        sheet["K" + str(i)] = '=IF(J' + str(i) + ' > '+ str(weight_limit-1) +', "Heavy", "Junior")'
        height = getHeight(payload)
        sheet["L" + str(i)] = height
        inches = getSize(height)
        if(gender == "M"):
            if(inches < s_size):
                sheet["M" + str(i)] = "Small"
            elif(inches < m_size):
                sheet["M" + str(i)] = "Medium"
            elif(inches < l_size):
                sheet["M" + str(i)] = "Large"
            elif(inches > l_size):
                sheet["M" + str(i)] = "Giant"
        elif(gender == "F"):
            sheet["M" + str(i)] = "Female"
        else:
            sheet["M" + str(i)] = "?"
        sheet["N" + str(i)] = getCountry(payload)
        sheet["O" + str(i)] = getDOB(payload)
        sheet["P" + str(i)] = getFinisher(payload)[0]
        sheet["Q" + str(i)] = getFinisher(payload)[1]
    wb.save(filename = xlsx_file)
        



    
